from mysql_orm import *
from common_lib import *
from flask import *
from sqlalchemy import func
from openpyxl import Workbook


history = Blueprint('history', __name__)


@history.route('/')
def render_history_page():
    return send_file('pages/history/history.html')


# TODO 获取device_id列表
@history.route('/get_all_device_id')
def get_all_device_id():
    try:
        user_cookie = request.cookies['LOGIN_SESSION']
        username = check_permission(cookie=user_cookie)
        if username:
            ret = db_get_all_device_id(username)
            print(ret)
            return api_resp(code=0, msg=ret)
        else:
            return api_resp_permission_error()
    except Exception as e:
        print(e)
        return api_resp_permission_error()


def db_get_all_device_id(username):
    username = str(username)
    ret_data = []
    data = db.session.query(device_rec.username, device_rec.device_id).group_by(device_rec.username, device_rec.device_id).having(device_rec.username == username)
    for i in data:
        ret_data.append({'username': i.username, 'device_id': i.device_id})
    return ret_data


# TODO 获取日期列表
@history.route('/get_date_options')
def get_date_options():
    try:
        user_cookie = request.cookies['LOGIN_SESSION']
        username = check_permission(cookie=user_cookie)
        if username:
            ret = db_get_date_options(username)
            print(ret)
            return api_resp(code=0, msg=ret)
        else:
            return api_resp_permission_error()
    except Exception as e:
        print(e)
        return api_resp_permission_error()


def db_get_date_options(username):
    username = str(username)
    ret_data = []
    date = db.session.query(func.date_format(device_rec.record_time, "%Y-%m-%d")).filter(device_rec.username == username).distinct()
    for i in date:
        ret_data.append(i[0])
    return ret_data


# TODO 根据device_id获取历史记录
@history.route('/get_history_data_by_device_id')
def get_history_data_by_device_id():
    try:
        user_cookie = request.cookies['LOGIN_SESSION']
        username = check_permission(cookie=user_cookie)
        if username:
            device_id = request.args.get("device_id")
            date = request.args.get("date")
            print('**', device_id, date)
            ret = db_get_history_data_by_device_id(username, device_id, date)
            return api_resp(code=0, msg=ret)
        else:
            return api_resp_permission_error()
    except Exception as e:
        print(e)
        return api_resp_permission_error()


def db_get_history_data_by_device_id(username, device_id, date):
    username = str(username)
    device_id = str(device_id)
    date = str(date)

    ret_data = []
    data = db.session.query(device_rec.record_type).filter(device_rec.username == username, device_rec.device_id == device_id).distinct()
    for i in data:
        tmp = {'device_id': device_id, 'record_type': i.record_type, 'history': None}
        history = []

        data_2 = db.session.query(device_rec.device_id, device_rec.record_type, device_rec.record_value, device_rec.unit, device_rec.record_time).filter(device_rec.username == username, device_rec.device_id == device_id, device_rec.record_type == i.record_type, device_rec.record_time.like('%' + date + '%'))
        for j in data_2:
            history.append({'record_value': j.record_value, 'record_time': str(j.record_time)[11:]})
            tmp['history'] = history
        ret_data.append(tmp)
    return ret_data


# TODO 导出历史记录
@history.route('/get_history_file')
def get_history_file():
    try:
        user_cookie = request.cookies['LOGIN_SESSION']
        username = check_permission(cookie=user_cookie)
        if username:
            device_id = request.args.get("device_id")
            date = request.args.get("date")
            ret = db_get_history_data_by_device_id(username, device_id, date)
            ret = dump_excel(ret, date)
            return api_resp(code=0, msg=ret)
        else:
            return api_resp_permission_error()
    except Exception as e:
        print(e)
        return api_resp_permission_error()


def dump_excel(data, date):
    wb = Workbook()
    ws = wb.active
    print(data)

    for i in data:
        # 写入时间
        ws.cell(row=1, column=1).value = str(date) + '时间'
        row_count = 2
        for j in i['history']:
            ws.cell(row=row_count, column=1).value = str(j['record_time'])
            row_count += 1

    # 写入值
    col_count = 2
    for i in data:
        row_count = 2
        ws.cell(row=1, column=col_count).value = str(i['device_id'] + '的' + i['record_type'])
        for j in i['history']:
            ws.cell(row=row_count, column=col_count).value = str(j['record_value'])
            row_count += 1
        col_count += 1

    wb.save('./dumps.xlsx')
    return '/history/download_history'


# TODO 下载文件
@history.route('/download_history')
def download_history():
    try:
        user_cookie = request.cookies['LOGIN_SESSION']
        username = check_permission(cookie=user_cookie)
        if username:
            return send_file('./dumps.xlsx')
        else:
            return api_resp_permission_error()
    except Exception as e:
        print(e)
        return api_resp_permission_error()
